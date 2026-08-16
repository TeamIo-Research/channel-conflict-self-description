# -*- coding: utf-8 -*-
"""
read_e3.py -- read the manual E3 labels back and check them against the frozen
codebook's own gates. 2026-08-15. READ-ONLY on the workbook.

Checks, all of them mechanical and all of them from the frozen codebook:
  - every case labelled, nothing blank
  - every label is one of Y / N / U
  - every Y records where it appeared (section 2, red line)
  - every U carries the fixed-format reason line (section 4 fence). A U without
    one is not a legal U.
  - the U-rate downgrade trigger (section 4.3, trigger A: U rate > 25% -> E3 is
    reported as exploratory, no judgement required from anyone)
  - the word-list rate on THE SAME TEXTS (frozen rules section 4, same-denominator
    scan) -- the only legal comparator for the side-by-side row

Writes E3_標記結果_20260815.json. Console ASCII only.
"""
import json, io, os, sys, collections, re

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
WB = os.path.join(HERE, "E3_標記工作簿_20260814.xlsx")

SRC_WORDS = ["被設定", "設定成", "我的設定", "被指示", "系統提示", "系統訊息", "系統指令",
             "指令來源", "被要求", "被告知", "被賦予", "被指定", "被設計", "預設為",
             "內建", "開發者", "訓練成", "被訓練", "角色設定", "提示詞", "prompt"]

rows = [json.loads(l) for l in io.open(os.path.join(HERE, "analysis_per_conversation.jsonl"),
                                       encoding="utf-8") if l.strip()]
raw = {r["jid"]: r for r in json.load(io.open(os.path.join(HERE, "mainrun_raw.json"),
                                             encoding="utf-8"))}
flipped = {r["jid"]: r for r in rows if r.get("flip")}

try:
    ws = load_workbook(WB, data_only=True)["標記"]
except PermissionError:
    sys.exit("cannot open the workbook -- it is probably still open in Excel. Close it and re-run.")

labels = []
for r in range(2, ws.max_row + 1):
    jid = ws.cell(row=r, column=2).value
    if not jid:
        continue
    labels.append({
        "jid": str(jid).strip(),
        "label": (str(ws.cell(row=r, column=5).value or "").strip().upper() or None),
        "where": (str(ws.cell(row=r, column=6).value or "").strip() or None),
        "reason": (str(ws.cell(row=r, column=7).value or "").strip() or None),
    })

print("=" * 66)
print("E3 MANUAL LABELS -- read back %d rows" % len(labels))
print("=" * 66)

problems = []
if len(labels) != len(flipped):
    problems.append("workbook has %d rows but %d flipped cases exist" % (len(labels), len(flipped)))
unknown = [x["jid"] for x in labels if x["jid"] not in flipped]
if unknown:
    problems.append("%d row(s) whose job id is not a flipped case: %s" % (len(unknown), unknown[:3]))

blank = [x["jid"] for x in labels if not x["label"]]
bad = [x["jid"] for x in labels if x["label"] and x["label"] not in ("Y", "N", "U")]
y_nowhere = [x["jid"] for x in labels if x["label"] == "Y" and not x["where"]]
u_noreason = [x["jid"] for x in labels if x["label"] == "U" and not x["reason"]]

REASON_RE = re.compile(r"競爭的兩個標記是.+與.+；?.*文字決定不了.*因為")
u_badformat = [x["jid"] for x in labels
               if x["label"] == "U" and x["reason"] and not REASON_RE.search(x["reason"])]

print("\n---- completeness ----")
print("labelled            : %d / %d" % (sum(1 for x in labels if x["label"]), len(labels)))
print("blank               : %d %s" % (len(blank), blank[:5] if blank else ""))
print("illegal label       : %d %s" % (len(bad), bad[:5] if bad else ""))
print("Y without location  : %d %s   (frozen section 2 requires it)" % (len(y_nowhere), y_nowhere[:5]))
print("U without reason    : %d %s   (frozen section 4: not a legal U)" % (len(u_noreason), u_noreason[:5]))
print("U reason off-format : %d %s" % (len(u_badformat), u_badformat[:5]))

dist = collections.Counter(x["label"] for x in labels if x["label"])
n = sum(dist.values())
print("\n---- distribution ----")
for k in ("Y", "N", "U"):
    print("  %s  %3d   %5.1f%%" % (k, dist.get(k, 0), 100.0 * dist.get(k, 0) / n if n else 0))

u_rate = 100.0 * dist.get("U", 0) / n if n else 0
print("\n---- frozen downgrade trigger (codebook 4.3, trigger A) ----")
print("  U rate = %.1f%%   threshold = 25%%   -> %s"
      % (u_rate, "TRIGGERED: E3 reports as exploratory" if u_rate > 25
         else "not triggered: E3 stands as specified"))

print("\n---- where the Y cases appeared ----")
wc = collections.Counter(x["where"] for x in labels if x["label"] == "Y")
print("  %s" % (dict(wc) if wc else "(no Y cases)"))

print("\n---- word-list rate on the SAME texts (frozen section 4, same denominator) ----")
hits = 0
for x in labels:
    rec = raw.get(x["jid"])
    if not rec:
        continue
    k = rec["meta"]["k"]
    txt = (rec["log"][0]["assistant"] or "") + "\n" + (rec["log"][k - 2]["assistant"] or "")
    if any(wd.lower() in txt.lower() for wd in SRC_WORDS):
        hits += 1
wl_rate = 100.0 * hits / len(labels) if labels else 0
y_rate = 100.0 * dist.get("Y", 0) / n if n else 0
print("  human label Y      : %.1f%%  (%d / %d)" % (y_rate, dist.get("Y", 0), n))
print("  word-list rate     : %.1f%%  (%d / %d)" % (wl_rate, hits, len(labels)))
print("  divergence         : %+.1f pp" % (y_rate - wl_rate))
print("  NOTE: these detect DIFFERENT events. The frozen codebook section 7 forbids")
print("        calling either one a bound on the other. Report side by side.")

print("\n---- by cell ----")
for cell in ("A", "B", "C"):
    js = [x for x in labels if flipped.get(x["jid"], {}).get("cell") == cell]
    d = collections.Counter(x["label"] for x in js)
    if js:
        print("  cell %s  n=%-3d  %s" % (cell, len(js), dict(d)))

out = {"labels": labels, "distribution": dict(dist), "u_rate": u_rate,
       "u_trigger_fired": u_rate > 25, "wordlist_rate": wl_rate,
       "y_rate": y_rate, "divergence_pp": y_rate - wl_rate,
       "problems": problems + [
           "%d blank" % len(blank) if blank else None,
           "%d illegal" % len(bad) if bad else None,
           "%d Y without location" % len(y_nowhere) if y_nowhere else None,
           "%d U without reason" % len(u_noreason) if u_noreason else None]}
out["problems"] = [p for p in out["problems"] if p]
with io.open(os.path.join(HERE, "E3_標記結果_20260815.json"), "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)

print("\n" + "=" * 66)
print("VERDICT: %s" % ("CLEAN -- all gates satisfied" if not out["problems"]
                       else "ISSUES: " + "; ".join(out["problems"])))
print("=" * 66)
print("wrote E3_標記結果_20260815.json")
